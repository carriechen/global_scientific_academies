from __future__ import annotations

import argparse
import time
import unicodedata
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote, urlparse

import requests
from openpyxl import load_workbook


DEFAULT_INPUT = Path(
    "global_science_academies V1_filled_ids.xlsx"
)
TARGET_COLUMNS = (
    "wikipedia_identifier",
    "Wikidata_identifier",
    "DBpedia_identifier",
)
USER_AGENT = "GSA-ID-Filler/1.0 (local research script)"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Fill missing Wikipedia, Wikidata, and DBpedia identifiers in an xlsx file."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Workbook to update.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output workbook path. Defaults to '<input stem>_entity_ids.xlsx'.",
    )
    parser.add_argument(
        "--sheet",
        help="Worksheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        help="Optional limit for testing on the first N data rows.",
    )
    return parser.parse_args()


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.casefold()
    for ch in "()[]{}|,:;.!?/'\"&-":
        text = text.replace(ch, " ")
    return " ".join(text.split())


def parse_entity_title(raw_url: Any) -> str | None:
    if is_blank(raw_url):
        return None

    parsed = urlparse(str(raw_url).strip())
    path = parsed.path.strip("/")
    if not path:
        return None

    title = path.rsplit("/", 1)[-1]
    return unquote(title).replace(" ", "_")


def build_wikipedia_url(title: str) -> str:
    return f"https://en.wikipedia.org/wiki/{quote(title, safe='_()')}"


def build_dbpedia_url(title: str) -> str:
    return f"https://dbpedia.org/page/{quote(title, safe='_()')}"


def safe_get(items: list[Any], index: int, default: Any = None) -> Any:
    if 0 <= index < len(items):
        return items[index]
    return default


class EntityResolver:
    def __init__(self) -> None:
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})
        self.qid_cache: dict[str, dict[str, str]] = {}
        self.wikipedia_cache: dict[str, dict[str, str]] = {}
        self.dbpedia_cache: dict[str, bool] = {}
        self.search_cache: dict[tuple[str, str], dict[str, str]] = {}

    def get_json(self, url: str, params: dict[str, Any]) -> dict[str, Any]:
        last_error: Exception | None = None
        for attempt in range(3):
            try:
                response = self.session.get(url, params=params, timeout=30)
                response.raise_for_status()
                return response.json()
            except Exception as exc:  # pragma: no cover - network variability
                last_error = exc
                time.sleep(1.0 * (attempt + 1))
        raise RuntimeError(f"Request failed for {url} with params={params}") from last_error

    def dbpedia_exists(self, url: str) -> bool:
        cached = self.dbpedia_cache.get(url)
        if cached is not None:
            return cached

        try:
            response = self.session.get(url, timeout=20, allow_redirects=True)
            exists = response.status_code == 200
        except Exception:
            exists = False

        self.dbpedia_cache[url] = exists
        return exists

    def fetch_from_qid(self, qid: str) -> dict[str, str]:
        if qid in self.qid_cache:
            return self.qid_cache[qid]

        data = self.get_json(
            "https://www.wikidata.org/w/api.php",
            {
                "action": "wbgetentities",
                "ids": qid,
                "props": "sitelinks",
                "sitefilter": "enwiki",
                "format": "json",
            },
        )
        entity = data.get("entities", {}).get(qid, {})
        enwiki = entity.get("sitelinks", {}).get("enwiki", {})
        title = enwiki.get("title")

        result = {"Wikidata_identifier": qid}
        if title:
            page_title = title.replace(" ", "_")
            result["wikipedia_identifier"] = build_wikipedia_url(page_title)
            dbpedia_url = build_dbpedia_url(page_title)
            if self.dbpedia_exists(dbpedia_url):
                result["DBpedia_identifier"] = dbpedia_url

        self.qid_cache[qid] = result
        return result

    def fetch_from_wikipedia(self, wikipedia_url: str) -> dict[str, str]:
        title = parse_entity_title(wikipedia_url)
        if not title:
            return {}
        if title in self.wikipedia_cache:
            return self.wikipedia_cache[title]

        data = self.get_json(
            "https://en.wikipedia.org/w/api.php",
            {
                "action": "query",
                "prop": "pageprops",
                "ppprop": "wikibase_item",
                "titles": title,
                "redirects": 1,
                "format": "json",
            },
        )
        pages = data.get("query", {}).get("pages", {})
        page = next(iter(pages.values()), {})
        qid = page.get("pageprops", {}).get("wikibase_item")

        result = {"wikipedia_identifier": build_wikipedia_url(title)}
        if qid:
            result.update(self.fetch_from_qid(qid))
            result["wikipedia_identifier"] = build_wikipedia_url(title)
        else:
            dbpedia_url = build_dbpedia_url(title)
            if self.dbpedia_exists(dbpedia_url):
                result["DBpedia_identifier"] = dbpedia_url

        self.wikipedia_cache[title] = result
        return result

    def fetch_from_dbpedia(self, dbpedia_url: str) -> dict[str, str]:
        title = parse_entity_title(dbpedia_url)
        if not title:
            return {}
        result = {}
        wikipedia_url = build_wikipedia_url(title)
        result["DBpedia_identifier"] = build_dbpedia_url(title)
        result.update(self.fetch_from_wikipedia(wikipedia_url))
        result["DBpedia_identifier"] = build_dbpedia_url(title)
        return result

    def search_wikidata(self, query: str) -> list[dict[str, Any]]:
        data = self.get_json(
            "https://www.wikidata.org/w/api.php",
            {
                "action": "wbsearchentities",
                "search": query,
                "language": "en",
                "format": "json",
                "limit": 10,
            },
        )
        return data.get("search", [])

    def score_candidate(
        self,
        item: dict[str, Any],
        normalized_names: list[str],
        country: str,
    ) -> int:
        label = normalize_text(item.get("label"))
        description = normalize_text(item.get("description"))
        aliases = [normalize_text(alias) for alias in item.get("aliases", [])]
        texts = [label, description, *aliases]
        score = 0

        for name in normalized_names:
            if not name:
                continue
            if label == name:
                score += 120
            elif name in label or label in name:
                score += 70

            for alias in aliases:
                if alias == name:
                    score += 100
                elif name in alias or alias in name:
                    score += 60

            name_tokens = set(name.split())
            for text in texts:
                tokens = set(text.split())
                overlap = len(name_tokens & tokens)
                score += overlap * 4

        if country and country in description:
            score += 25
        if "academy" in description or "academy" in label:
            score += 8
        if "science" in description or "technology" in description:
            score += 4
        if item.get("id", "").startswith("Q"):
            score += 1
        return score

    def resolve_from_search(self, row_values: dict[str, Any]) -> dict[str, str]:
        acad_name_en = str(row_values.get("acad_name_en") or "").strip()
        acad_name = str(row_values.get("acad_name") or "").strip()
        acad_abbr = str(row_values.get("acad_name_abbreviation") or "").strip()
        country_raw = str(row_values.get("country") or "").strip()
        country = normalize_text(country_raw)

        query_parts = []
        for value in (acad_name_en, acad_name, acad_abbr):
            if value and value not in query_parts:
                query_parts.append(value)

        normalized_names = [normalize_text(name) for name in query_parts if name]
        best_item: dict[str, Any] | None = None
        best_score = 0

        for query in query_parts:
            cache_key = (query, country_raw)
            if cache_key in self.search_cache:
                cached = self.search_cache[cache_key]
                if cached:
                    return cached
                continue

            search_terms = [query]
            if country_raw:
                search_terms.append(f"{query} {country_raw}")

            candidates: list[dict[str, Any]] = []
            seen_ids: set[str] = set()
            for term in search_terms:
                for item in self.search_wikidata(term):
                    qid = item.get("id")
                    if qid and qid not in seen_ids:
                        seen_ids.add(qid)
                        candidates.append(item)

            for item in candidates:
                score = self.score_candidate(item, normalized_names, country)
                if score > best_score:
                    best_score = score
                    best_item = item

            self.search_cache[cache_key] = {}

        if not best_item or best_score < 80:
            return {}

        resolved = self.fetch_from_qid(best_item["id"])
        for query in query_parts:
            self.search_cache[(query, country_raw)] = resolved
        return resolved


def fill_missing_identifiers(
    input_path: Path,
    output_path: Path,
    sheet_name: str | None,
    max_rows: int | None,
) -> tuple[int, int, list[str]]:
    workbook = load_workbook(input_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in worksheet[1]]
    column_map = {header: idx + 1 for idx, header in enumerate(headers)}

    required = {
        "acad_name",
        "acad_name_abbreviation",
        "acad_name_en",
        "country",
        *TARGET_COLUMNS,
    }
    missing_headers = sorted(required - set(column_map))
    if missing_headers:
        raise ValueError(f"Missing required columns: {', '.join(missing_headers)}")

    resolver = EntityResolver()
    updated_cells = 0
    touched_rows = 0
    unresolved_rows: list[str] = []

    for excel_row_idx in range(2, worksheet.max_row + 1):
        if max_rows is not None and (excel_row_idx - 1) > max_rows:
            break

        row_values = {
            header: worksheet.cell(row=excel_row_idx, column=col_idx).value
            for header, col_idx in column_map.items()
        }
        if not any(is_blank(row_values[col]) for col in TARGET_COLUMNS):
            continue

        resolved: dict[str, str] = {}
        if not is_blank(row_values["Wikidata_identifier"]):
            resolved.update(
                resolver.fetch_from_qid(str(row_values["Wikidata_identifier"]).strip())
            )
        if not is_blank(row_values["wikipedia_identifier"]):
            resolved.update(
                resolver.fetch_from_wikipedia(str(row_values["wikipedia_identifier"]).strip())
            )
        if not is_blank(row_values["DBpedia_identifier"]):
            resolved.update(
                resolver.fetch_from_dbpedia(str(row_values["DBpedia_identifier"]).strip())
            )
        if not resolved:
            resolved.update(resolver.resolve_from_search(row_values))

        row_updated = False
        for column_name in TARGET_COLUMNS:
            if is_blank(row_values[column_name]) and resolved.get(column_name):
                worksheet.cell(
                    row=excel_row_idx,
                    column=column_map[column_name],
                    value=resolved[column_name],
                )
                updated_cells += 1
                row_updated = True

        if row_updated:
            touched_rows += 1
        elif any(is_blank(row_values[col]) for col in TARGET_COLUMNS):
            identifier = row_values.get("acad_id") or row_values.get("acad_name_en") or excel_row_idx
            unresolved_rows.append(str(identifier))

    workbook.save(output_path)
    return touched_rows, updated_cells, unresolved_rows


def main() -> None:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else input_path.with_name(f"{input_path.stem}_entity_ids.xlsx")
    )

    touched_rows, updated_cells, unresolved_rows = fill_missing_identifiers(
        input_path=input_path,
        output_path=output_path,
        sheet_name=args.sheet,
        max_rows=args.max_rows,
    )

    print(f"Saved updated workbook to: {output_path}")
    print(f"Rows updated: {touched_rows}")
    print(f"Cells filled: {updated_cells}")
    print(f"Rows still unresolved: {len(unresolved_rows)}")
    if unresolved_rows:
        print("Unresolved row identifiers:")
        for item in unresolved_rows:
            print(f"- {item}")


if __name__ == "__main__":
    main()
